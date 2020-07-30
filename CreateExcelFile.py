import openpyxl

class ExcelUtility(object):

    def __init__(self):
        print("Read write to excel file")

    def group(self, lst, n):
        """group([0,3,4,10,2,3], 2) => [(0,3), (4,10), (2,3)]

        Group a list into consecutive n-tuples. Incomplete tuples are
        discarded e.g.

        >>> group(range(10), 3)
        [(0, 1, 2), (3, 4, 5), (6, 7, 8)]
        """
        return zip(*[lst[i::n] for i in range(n)])

    def Total_Rows_From_Excel(self, filename, sheetname):
        # open excel file
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[sheetname]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        return totalrows

    def Total_Columns_From_Excel(self, filename, sheetname):
        # open excel file
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[sheetname]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        return totalcolumns

    def Replace_HashTag(self, teststring):
        return teststring.replace("#", "")

    def Return_Basket_URL(self, teststring1):
        firstval1, firstval2 = teststring1.split('=', 1)
        return firstval2.replace("#/", "")

    def Read_Data_From_Excel(self, filename, sheetname, testcaseid, columnName):
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[sheetname]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        ColumnNumber = None

        for colnum in range(1, totalcolumns + 1):
            header = worksheet.cell(row=1, column=colnum).value
            if header == columnName:
                ColumnNumber = colnum
                break

        for rownum in range(2, totalrows + 1):
            TestCaseNumber = worksheet.cell(row=rownum, column=1).value
            if str(TestCaseNumber) == str(testcaseid):
                searchData = worksheet.cell(rownum, ColumnNumber).value
                print(searchData)
                return searchData

    def Write_Data_Into_Excel(self, filename, sheetname, testcaseid, columnName, writenewvalue):
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[sheetname]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        ColumnNumber = None

        for colnum in range(1, totalcolumns + 1):
            header = worksheet.cell(row=1, column=colnum).value
            if header == columnName:
                ColumnNumber = colnum
                break

        for rownum in range(2, totalrows + 1):
            TestCaseNumber = worksheet.cell(row=rownum, column=1).value
            if str(TestCaseNumber) == str(testcaseid):
                searchData = worksheet.cell(rownum, ColumnNumber)
                searchData.value = writenewvalue
                workbook.save(filename)
                break

    def Write_Data_Into_Excel_Into_NewRow(self, filename, SheetName, RowNumber, columnName, writenewvalue):
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[SheetName]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        ColumnNumber = None

        for colnum in range(1, totalcolumns + 1):
            header = worksheet.cell(row=1, column=colnum).value
            if header == columnName:
                ColumnNumber = colnum
                break

        searchData = worksheet.cell(RowNumber+1, ColumnNumber)
        searchData.value = writenewvalue
        workbook.save(filename)

    def Read_All_Active_Subscriptions_From_Excel(self, filename, Sheetname, testcasestatus, columnName):
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[Sheetname]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        ColumnNumber = None
        my_list = []

        for colnum in range(1, totalcolumns + 1):
            header = worksheet.cell(row=1, column=colnum).value
            if header == columnName:
                ColumnNumber = colnum
                break

        for rownum in range(totalrows, 2, -1):
            for colnum in range(1, totalcolumns + 1):
                TestCaseNumber = worksheet.cell(row=rownum, column=colnum).value
                if str(TestCaseNumber) == str(testcasestatus):
                    searchData = worksheet.cell(row=rownum, column=ColumnNumber).value
                    if searchData != '':
                        my_list.append(searchData)
                        break
        return my_list

    def Read_Data_From_Excel_Row(self, filename, Sheetname, rownumber, columnName):
        workbook = openpyxl.load_workbook(filename)
        # open work sheet
        worksheet = workbook[Sheetname]
        #
        # worksheet.cell(3,2).value
        # totalrows = worksheet.row_values(-1)
        # totalcolumns = worksheet.col_values(-1)
        totalrows = worksheet.max_row
        totalcolumns = worksheet.max_column
        ColumnNumber = None

        for colnum in range(1, totalcolumns + 1):
            header = worksheet.cell(row=1, column=colnum).value
            if header == columnName:
                ColumnNumber = colnum
                break

        searchData = worksheet.cell(row=rownumber + 1, column=ColumnNumber).value
        return searchData